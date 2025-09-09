#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
재발송 변환기 - Streamlit 웹앱
통합수집기로 추출한 데이터를 '수기_재발송양식'에 맞게 변환하는 웹앱
클레임유형이 '교환' 또는 '해당없음'인 데이터만 변환
"""

import streamlit as st
import pandas as pd
import openpyxl
import io
from datetime import datetime
import tempfile
import os

# 페이지 설정
st.set_page_config(
    page_title="재발송 변환기",
    page_icon="📦",
    layout="centered"
)

# --- 헬퍼 함수 ---
def get_current_date_string():
    """현재 날짜를 YYYYMMDD 형식으로 반환"""
    return datetime.now().strftime('%Y%m%d')

def generate_bundle_numbers(df):
    """주소별로 묶음배송번호를 생성"""
    bundle_numbers = {}
    current_time = datetime.now()
    time_str = current_time.strftime('%Y%m%d%H%M')
    
    # 주소별로 그룹화
    unique_addresses = df['주소'].dropna().unique()
    
    # 주소별로 고유 번호 부여 (같은 주소는 같은 번호)
    for i, address in enumerate(unique_addresses, 1):
        # 시간 부분은 고정하고, 주소별로 고유 번호 부여
        bundle_number = f"re{time_str}{i:02d}"
        bundle_numbers[str(address)] = bundle_number
    
    return bundle_numbers

# --- 데이터 변환 함수 ---
def convert_data_to_reshipment(df, bundle_numbers=None):
    """통합수집기 데이터를 재발송 양식으로 변환 (교환 또는 해당없음만)"""
    # 클레임유형이 '교환' 또는 '해당없음'인 데이터만 필터링
    filtered_df = df[df.get('클레임유형', '').isin(['교환', '해당없음'])].copy()
    
    if len(filtered_df) == 0:
        return pd.DataFrame()
    
    # 묶음배송번호가 제공되지 않으면 생성
    if bundle_numbers is None:
        bundle_numbers = generate_bundle_numbers(filtered_df)
    
    converted_data = []
    for idx, row in filtered_df.iterrows():
        address = str(row.get('주소', ''))
        bundle_number = bundle_numbers.get(address, '')
        
        converted_row = {
            '품목코드': str(row.get('품목코드', '')),
            '가격': str(row.get('총결제금액', '')),
            '품목수량': str(row.get('주문수량', '')),
            '받는사람명': str(row.get('주문자명', '')),
            '받는사람 전화번호': str(row.get('연락처', '')),
            '받는사람 우편번호': str(row.get('우편번호', '')),
            '받는사람 주소': str(row.get('주소', '')),
            '묶음배송번호': bundle_number,
            '주문일자': get_current_date_string()
        }
        converted_data.append(converted_row)
    
    return pd.DataFrame(converted_data)

def create_excel_file(converted_df):
    """재발송 양식 엑셀 파일 생성 (하드코딩된 템플릿 구조)"""
    # 메모리에서 엑셀 파일 생성
    output = io.BytesIO()
    
    # 새 워크북 생성
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # 원본 템플릿의 정확한 헤더 (31개 칼럼)
    headers = [
        '* F/C', '* 주문유형', '* 배송처', '* 고객ID', '판매채널', '* 묶음배송번호', '* 품목코드', 
        '품목명', '옵션', '가격', '* 품목수량', '주문자', '* 받는사람명', '주문자 전화번호', 
        '* 받는사람 전화번호', '* 받는사람 우편번호', '* 받는사람 주소', '배송메세지', '* 주문일자', 
        '상품주문번호', '주문번호(참조)', '주문중개채널(상세)', '박스구분', '상세배송유형', 
        '새벽배송 SMS 전송', '새벽배송 현관비밀번호', '위험물 구분', '* 주문중개채널', 
        'API 연동용 판매자ID', '* 주문시간', '받는사람 핸드폰'
    ]
    
    # 헤더 입력 (1행)
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # 데이터 입력 (2행부터)
    start_row = 2
    for idx, row in converted_df.iterrows():
        current_row = start_row + idx
        
        # 우편번호 5자리 고정 처리
        postal_code = str(row.get('받는사람 우편번호', '')).strip()
        if postal_code and len(postal_code) == 4:
            postal_code = '0' + postal_code  # 4자리면 앞에 0 추가
        elif not postal_code:
            postal_code = '00000'  # 빈 값이면 00000
        
        # 필수고정값들 (모든 행에 동일하게 입력)
        worksheet.cell(row=current_row, column=1, value="NS001")      # * F/C
        worksheet.cell(row=current_row, column=2, value="7")          # * 주문유형
        worksheet.cell(row=current_row, column=3, value="17")         # * 배송처
        worksheet.cell(row=current_row, column=4, value="90015746")   # * 고객ID
        worksheet.cell(row=current_row, column=5, value="NFA")        # 판매채널
        worksheet.cell(row=current_row, column=6, value=str(row.get('묶음배송번호', '')))  # * 묶음배송번호
        worksheet.cell(row=current_row, column=7, value=str(row.get('품목코드', '')))     # * 품목코드
        worksheet.cell(row=current_row, column=8, value="")           # 품목명
        worksheet.cell(row=current_row, column=9, value="")           # 옵션
        worksheet.cell(row=current_row, column=10, value=str(row.get('가격', '')))       # 가격
        worksheet.cell(row=current_row, column=11, value=str(row.get('품목수량', '')))   # * 품목수량
        worksheet.cell(row=current_row, column=12, value="")          # 주문자
        worksheet.cell(row=current_row, column=13, value=str(row.get('받는사람명', ''))) # * 받는사람명
        worksheet.cell(row=current_row, column=14, value="")          # 주문자 전화번호
        worksheet.cell(row=current_row, column=15, value=str(row.get('받는사람 전화번호', ''))) # * 받는사람 전화번호
        worksheet.cell(row=current_row, column=16, value=postal_code) # * 받는사람 우편번호
        worksheet.cell(row=current_row, column=17, value=str(row.get('받는사람 주소', ''))) # * 받는사람 주소
        worksheet.cell(row=current_row, column=18, value="")          # 배송메세지
        worksheet.cell(row=current_row, column=19, value=str(row.get('주문일자', '')))   # * 주문일자
        worksheet.cell(row=current_row, column=20, value="")          # 상품주문번호
        worksheet.cell(row=current_row, column=21, value="")          # 주문번호(참조)
        worksheet.cell(row=current_row, column=22, value="")          # 주문중개채널(상세)
        worksheet.cell(row=current_row, column=23, value="")          # 박스구분
        worksheet.cell(row=current_row, column=24, value="")          # 상세배송유형
        worksheet.cell(row=current_row, column=25, value="")          # 새벽배송 SMS 전송
        worksheet.cell(row=current_row, column=26, value="")          # 새벽배송 현관비밀번호
        worksheet.cell(row=current_row, column=27, value="")          # 위험물 구분
        worksheet.cell(row=current_row, column=28, value="SELF")      # * 주문중개채널
        worksheet.cell(row=current_row, column=29, value="")          # API 연동용 판매자ID
        worksheet.cell(row=current_row, column=30, value="09:00:00")  # * 주문시간
        worksheet.cell(row=current_row, column=31, value="")          # 받는사람 핸드폰
    
    # 파일 저장
    workbook.save(output)
    output.seek(0)
    return output

# --- 메인 앱 ---
def main():
    # 컨테이너로 가운데 정렬
    with st.container():
        # 제목
        st.title("📦 재발송 변환기")
        st.markdown("---")
        
        # 설명
        st.markdown("통합수집기로 추출한 데이터를 재발송 양식으로 변환합니다.")
        
        # 파일 업로드
        uploaded_file = st.file_uploader(
            "엑셀 파일을 업로드하세요",
            type=['xlsx']
        )
    
        if uploaded_file is not None:
            try:
                # 파일 읽기
                df = pd.read_excel(uploaded_file)
                
                # 변환 버튼
                if st.button("🔄 변환 시작"):
                    with st.spinner("변환 중..."):
                        # 묶음배송번호 생성
                        bundle_numbers = generate_bundle_numbers(df)
                        
                        # 데이터 변환
                        converted_df = convert_data_to_reshipment(df, bundle_numbers)
                        
                        if len(converted_df) > 0:
                            # 변환 결과 표시
                            st.success(f"✅ 변환 완료! 총 {len(converted_df)}행이 변환되었습니다.")
                            
                            # 변환된 데이터 미리보기
                            st.markdown("### 📋 변환된 데이터 미리보기")
                            st.dataframe(converted_df, use_container_width=True)
                            
                            # 엑셀 파일 생성
                            excel_file = create_excel_file(converted_df)
                            
                            # 다운로드 버튼
                            st.download_button(
                                label="📥 다운로드",
                                data=excel_file.getvalue(),
                                file_name=f"수기_재발송양식_변환결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                                
                        else:
                            st.warning("⚠️ 변환할 데이터가 없습니다. 클레임유형이 '교환' 또는 '해당없음'인 데이터가 있는지 확인해주세요.")
                            
            except Exception as e:
                st.error(f"❌ 파일 처리 중 오류가 발생했습니다: {str(e)}")
    

if __name__ == "__main__":
    main()
